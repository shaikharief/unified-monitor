#!/usr/bin/env python3
"""
3-Month Plan Order Monitor v1.1
================================
Polls Databricks mvno_order table for new orders matching:
  - "3 Month Plan"
  - "Intro Price - 3 Months"

Writes order details to a formatted Excel (.xlsx) file and sends Slack alerts.

Usage:
  python3 three_month_monitor.py                     # Run continuous monitor (polls every 5 min)
  python3 three_month_monitor.py --once              # Single check and exit
  python3 three_month_monitor.py --interval 120      # Poll every 2 minutes
  python3 three_month_monitor.py --no-slack          # Disable Slack alerts
"""

import os
import sys
import ssl
import json
import time
import urllib.request
import urllib.error
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl not found. Install: pip install openpyxl")

# ══════════════════════════════════════════════════════════════════════
#   CONFIG
# ══════════════════════════════════════════════════════════════════════

DATABRICKS = {
    "host": os.environ.get("DATABRICKS_HOST", "dbc-b7af8d94-a7ba.cloud.databricks.com"),
    "token": os.environ.get("DATABRICKS_TOKEN", ""),
    "warehouse_id": os.environ.get("DATABRICKS_WAREHOUSE", "7352f0009fd4e049"),
    "catalog": "rds-prod_catalog",
    "schema": "cj_prod",
}

SLACK_WEBHOOK_URL = os.environ.get("SLACK_WEBHOOK", "")

WATCH_PRODUCTS = [
    "3 Month Plan",
    "Intro Price - 3 Months",
]

POLL_INTERVAL = 300
IST = timezone(timedelta(hours=5, minutes=30))

# ══════════════════════════════════════════════════════════════════════
#   EXCEL COLUMN DEFINITIONS
# ══════════════════════════════════════════════════════════════════════

COLUMNS = [
    {"key": "order_id",           "header": "Order ID",          "width": 12},
    {"key": "customer_name",      "header": "Customer",          "width": 22},
    {"key": "email",              "header": "Email",             "width": 30},
    {"key": "phone_number",       "header": "Phone",             "width": 18},
    {"key": "product_name",       "header": "Plan",              "width": 24},
    {"key": "amount",             "header": "Amount ($)",        "width": 14},
    {"key": "select_number_type", "header": "New / Port-In",     "width": 16},
    {"key": "portin_status",      "header": "Port-In Status",    "width": 16},
    {"key": "payment_status",     "header": "Payment Status",    "width": 34},
    {"key": "order_status",       "header": "Order Status",      "width": 16},
    {"key": "esim_status",        "header": "eSIM Status",       "width": 14},
    {"key": "payment_intent_id",  "header": "Payment Intent",    "width": 32},
    {"key": "billing_account_id", "header": "Billing Account",   "width": 28},
    {"key": "individual_id",      "header": "Individual ID",     "width": 28},
    {"key": "customer_id",        "header": "Customer ID",       "width": 28},
    {"key": "redemption_code",    "header": "Redemption Code",   "width": 18},
    {"key": "created_at",         "header": "Created At (IST)",  "width": 24},
    {"key": "recorded_at",        "header": "Recorded At (IST)", "width": 24},
]

# ══════════════════════════════════════════════════════════════════════
#   STYLES
# ══════════════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGN = Alignment(vertical="center", wrap_text=False)
MONEY_ALIGN = Alignment(horizontal="right", vertical="center")

ROW_FILL_EVEN = PatternFill("solid", fgColor="F2F7FB")
ROW_FILL_ODD = PatternFill("solid", fgColor="FFFFFF")

GREEN_FONT = Font(name="Arial", size=10, color="006100")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FONT = Font(name="Arial", size=10, color="9C0006")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
AMBER_FONT = Font(name="Arial", size=10, color="9C6500")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2EC"),
    right=Side(style="thin", color="D9E2EC"),
    top=Side(style="thin", color="D9E2EC"),
    bottom=Side(style="thin", color="D9E2EC"),
)

TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Arial", size=10, color="666666")

# ══════════════════════════════════════════════════════════════════════
#   HELPERS
# ══════════════════════════════════════════════════════════════════════

def now_utc():
    return datetime.now(timezone.utc)

def fmt_ist(dt):
    if dt is None:
        return "-"
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
        except Exception:
            return dt
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(IST).strftime("%Y-%m-%d %I:%M:%S %p")

def log(msg, level="INFO"):
    ts = fmt_ist(now_utc())
    print(f"[{ts}] [{level}] {msg}")

# ══════════════════════════════════════════════════════════════════════
#   DATABRICKS CLIENT
# ══════════════════════════════════════════════════════════════════════

def databricks_query(sql):
    url = f"https://{DATABRICKS['host']}/api/2.0/sql/statements/"
    payload = json.dumps({
        "warehouse_id": DATABRICKS["warehouse_id"],
        "statement": sql,
        "wait_timeout": "30s",
        "disposition": "INLINE",
        "format": "JSON_ARRAY",
    }).encode("utf-8")
    headers = {
        "Authorization": f"Bearer {DATABRICKS['token']}",
        "Content-Type": "application/json",
    }
    ctx = ssl.create_default_context()
    req = urllib.request.Request(url, data=payload, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        log(f"Databricks query failed: {e}", "ERROR")
        return []

    if data.get("status", {}).get("state") != "SUCCEEDED":
        log(f"Query status: {data.get('status', {}).get('state')}", "ERROR")
        return []

    columns = [c["name"] for c in data.get("manifest", {}).get("schema", {}).get("columns", [])]
    raw_rows = data.get("result", {}).get("data_array", [])
    rows = []
    for raw_row in raw_rows:
        row = {}
        if isinstance(raw_row, list):
            for i, col in enumerate(columns):
                if i < len(raw_row):
                    row[col] = raw_row[i] if raw_row[i] is not None else None
                else:
                    row[col] = None
        elif isinstance(raw_row, dict):
            values = raw_row.get("values", [])
            for i, col in enumerate(columns):
                if i < len(values):
                    v = values[i]
                    if isinstance(v, dict):
                        row[col] = v.get("string_value") if "string_value" in v else None
                    else:
                        row[col] = v if v is not None else None
                else:
                    row[col] = None
        else:
            continue
        rows.append(row)
    return rows

# ══════════════════════════════════════════════════════════════════════
#   FETCH 3-MONTH PLAN ORDERS
# ══════════════════════════════════════════════════════════════════════

def fetch_three_month_orders():
    cat, sch = DATABRICKS["catalog"], DATABRICKS["schema"]
    conditions = " OR ".join(f"LOWER(o.product_name) LIKE '%{p.lower()}%'" for p in WATCH_PRODUCTS)
    sql = f"""
    SELECT
        o.order_id, o.status AS order_status, o.product_name,
        o.amount, o.currency, o.select_number_type, o.portin_status,
        o.payment_info, o.payment_intent_id, o.billing_account_id,
        o.individual_id, o.user_id, o.esim_status,
        o.created_at, o.updated_at,
        o.first_name AS order_first_name, o.last_name AS order_last_name,
        o.redemption_code,
        c.user_name AS email, c.given_name, c.family_name,
        c.phone_number, c.customer_id
    FROM `{cat}`.`{sch}`.mvno_order o
    LEFT JOIN `{cat}`.`{sch}`.certification c ON o.user_id = c.user_id
    WHERE ({conditions})
    ORDER BY o.created_at DESC
    """
    return databricks_query(sql)

# ══════════════════════════════════════════════════════════════════════
#   PAYMENT STATUS HELPER
# ══════════════════════════════════════════════════════════════════════

def get_payment_status(row):
    payment_info = row.get("payment_info")
    payment_intent = row.get("payment_intent_id")
    if payment_info and payment_info.strip() not in ("", "null", "None", "[]"):
        try:
            pi_data = json.loads(payment_info)
            if isinstance(pi_data, list) and pi_data:
                aid = pi_data[0].get("id", "")[:24]
                return f"CONFIRMED ({aid})"
        except Exception:
            pass
        return "CONFIRMED"
    elif payment_intent and payment_intent.strip():
        return f"PENDING ({payment_intent[:30]})"
    return "NONE"

# ══════════════════════════════════════════════════════════════════════
#   EXCEL FILE MANAGEMENT
# ══════════════════════════════════════════════════════════════════════

def create_workbook(filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "3-Month Plan Orders"

    ws.merge_cells("A1:H1")
    ws["A1"] = "3-Month Plan Orders - Live Tracker"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Auto-generated by 3-Month Plan Monitor v1.1 | Watching: {', '.join(WATCH_PRODUCTS)}"
    ws["A2"].font = SUBTITLE_FONT
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    header_row = 4
    for col_idx, col_def in enumerate(COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_def["header"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]
    ws.row_dimensions[header_row].height = 30

    ws.freeze_panes = "A5"
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A4:{last_col}4"

    wb.save(filepath)
    log(f"Created Excel file: {filepath}")
    return wb

def get_recorded_order_ids(filepath):
    if not filepath.exists():
        return set()
    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        ids = set()
        for row in ws.iter_rows(min_row=5, max_col=1, values_only=True):
            if row[0]:
                ids.add(str(row[0]))
        wb.close()
        return ids
    except Exception:
        return set()

def get_next_data_row(ws):
    for row_num in range(5, ws.max_row + 2):
        if ws.cell(row=row_num, column=1).value is None:
            return row_num
    return ws.max_row + 1

def apply_status_style(cell, value):
    if value is None:
        return
    val_upper = str(value).upper()
    if val_upper in ("COMPLETED", "DONE", "NEW_NUMBER") or "CONFIRMED" in val_upper:
        cell.font = GREEN_FONT
        cell.fill = GREEN_FILL
    elif val_upper in ("FAILED", "CANCELLED") or val_upper == "NONE":
        cell.font = RED_FONT
        cell.fill = RED_FILL
    elif val_upper in ("INPROGRESS", "PENDING", "DRAFT", "PORT_IN", "PORTIN") or "PENDING" in val_upper:
        cell.font = AMBER_FONT
        cell.fill = AMBER_FILL

def append_order_to_excel(filepath, order):
    wb = load_workbook(filepath)
    ws = wb.active
    row_num = get_next_data_row(ws)
    is_even = (row_num - 4) % 2 == 0
    bg_fill = ROW_FILL_EVEN if is_even else ROW_FILL_ODD

    for col_idx, col_def in enumerate(COLUMNS, 1):
        key = col_def["key"]
        value = order.get(key, "-")
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGN
        cell.border = THIN_BORDER
        cell.fill = bg_fill

        if key == "amount":
            cell.alignment = MONEY_ALIGN
            cell.number_format = '$#,##0.00'
            try:
                cell.value = float(value) if value and value != "-" else 0
            except (ValueError, TypeError):
                pass

        if key in ("payment_status", "order_status", "select_number_type", "esim_status"):
            apply_status_style(cell, value)

    ws.row_dimensions[row_num].height = 22

    total_orders = row_num - 4
    ws["A2"] = f"Auto-generated by 3-Month Plan Monitor v1.1 | Total orders: {total_orders} | Last updated: {fmt_ist(now_utc())}"
    ws["A2"].font = SUBTITLE_FONT

    wb.save(filepath)

def update_orders_in_excel(filepath, updated_orders):
    if not updated_orders:
        return
    wb = load_workbook(filepath)
    ws = wb.active
    changed = False
    for row_num in range(5, ws.max_row + 1):
        oid = str(ws.cell(row=row_num, column=1).value or "")
        if oid in updated_orders:
            upd = updated_orders[oid]
            for col_idx, col_def in enumerate(COLUMNS, 1):
                key = col_def["key"]
                if key in ("order_status", "payment_status", "esim_status"):
                    new_val = upd.get(key)
                    if new_val:
                        cell = ws.cell(row=row_num, column=col_idx, value=new_val)
                        cell.font = DATA_FONT
                        cell.border = THIN_BORDER
                        apply_status_style(cell, new_val)
                        changed = True
    if changed:
        wb.save(filepath)

# ══════════════════════════════════════════════════════════════════════
#   SLACK NOTIFICATION
# ══════════════════════════════════════════════════════════════════════

def send_slack(order):
    if not SLACK_WEBHOOK_URL:
        return

    status_emoji = {"COMPLETED": "✅", "INPROGRESS": "🔄", "PENDING": "⏳", "DRAFT": "📝"}.get(
        (order.get("order_status") or "").upper(), "❓")
    pay_status = order.get("payment_status", "NONE")
    pay_emoji = "✅" if "CONFIRMED" in pay_status else ("⏳" if "PENDING" in pay_status else "❌")
    num_type = order.get("select_number_type", "-")
    num_emoji = "🆕" if num_type == "NEW_NUMBER" else ("🔄" if num_type in ("PORT_IN", "PORTIN") else "❓")

    blocks = [
        {"type": "header", "text": {"type": "plain_text", "text": "📦 NEW 3-MONTH PLAN ORDER", "emoji": True}},
        {"type": "section", "text": {"type": "mrkdwn", "text": (
            f"*Order #{order['order_id']}* | {status_emoji} {order.get('order_status', '-')}\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
        )}},
        {"type": "section", "fields": [
            {"type": "mrkdwn", "text": f"👤 *Customer*\n{order.get('customer_name', '-')}"},
            {"type": "mrkdwn", "text": f"📧 *Email*\n{order.get('email', '-')}"},
            {"type": "mrkdwn", "text": f"📱 *Phone*\n{order.get('phone_number', '-') or '-'}"},
            {"type": "mrkdwn", "text": f"📋 *Plan*\n{order.get('product_name', '-')}"},
            {"type": "mrkdwn", "text": f"💰 *Amount*\n${order.get('amount', '?')} {order.get('currency', 'USD')}"},
            {"type": "mrkdwn", "text": f"{num_emoji} *Number Type*\n{num_type}"},
            {"type": "mrkdwn", "text": f"{pay_emoji} *Payment*\n{pay_status}"},
            {"type": "mrkdwn", "text": f"🕐 *Created*\n{order.get('created_at', '-')}"},
        ]},
        {"type": "divider"},
        {"type": "context", "elements": [
            {"type": "mrkdwn", "text": (
                f"_Billing: {order.get('billing_account_id', '-')[:24]} | "
                f"Individual: {order.get('individual_id', '-')[:24]} | "
                f"3-Month Plan Monitor v1.1_"
            )}
        ]},
    ]

    payload = json.dumps({"channel": "#order-alerts", "blocks": blocks}).encode("utf-8")
    req = urllib.request.Request(
        SLACK_WEBHOOK_URL, data=payload,
        headers={"Content-Type": "application/json"}, method="POST",
    )
    ctx = ssl.create_default_context()
    try:
        with urllib.request.urlopen(req, timeout=10, context=ctx) as resp:
            if resp.status == 200:
                log(f"  ✅ Slack alert sent for Order #{order['order_id']}")
            else:
                log(f"  ❌ Slack HTTP {resp.status}", "ERROR")
    except Exception as e:
        log(f"  ❌ Slack error: {e}", "ERROR")

# ══════════════════════════════════════════════════════════════════════
#   MAIN POLL LOOP
# ══════════════════════════════════════════════════════════════════════

def build_order_dict(row):
    gn = row.get("given_name") or row.get("order_first_name") or ""
    fn = row.get("family_name") or row.get("order_last_name") or ""
    customer_name = f"{gn} {fn}".strip() or "-"
    return {
        "order_id": row.get("order_id", "-"),
        "customer_name": customer_name,
        "email": row.get("email", "-"),
        "phone_number": row.get("phone_number") or "-",
        "product_name": row.get("product_name", "-"),
        "amount": row.get("amount", "0"),
        "currency": row.get("currency", "USD"),
        "select_number_type": row.get("select_number_type", "-"),
        "portin_status": row.get("portin_status", "-"),
        "payment_status": get_payment_status(row),
        "payment_intent_id": row.get("payment_intent_id") or "-",
        "order_status": row.get("order_status", "-"),
        "esim_status": row.get("esim_status") or "-",
        "billing_account_id": row.get("billing_account_id") or "-",
        "individual_id": row.get("individual_id") or "-",
        "customer_id": row.get("customer_id") or "-",
        "redemption_code": row.get("redemption_code") or "-",
        "created_at": fmt_ist(row.get("created_at")),
        "recorded_at": fmt_ist(now_utc()),
    }

def poll(output_file, send_alerts=True):
    log("Polling Databricks for 3-month plan orders...")
    rows = fetch_three_month_orders()
    if not rows:
        log("  No 3-month plan orders found yet.")
        return 0

    log(f"  Found {len(rows)} total 3-month plan order(s) in database.")
    recorded_ids = get_recorded_order_ids(output_file)
    new_count = 0
    updated_orders = {}

    for row in rows:
        oid = row.get("order_id", "")
        order = build_order_dict(row)

        if oid not in recorded_ids:
            append_order_to_excel(output_file, order)
            new_count += 1
            log(f"  🆕 NEW: Order #{oid} | {order['product_name']} | "
                f"${order['amount']} | {order['customer_name']} | {order['order_status']}")
            if send_alerts:
                send_slack(order)
        else:
            updated_orders[oid] = order

    if updated_orders:
        update_orders_in_excel(output_file, updated_orders)
        log(f"  Updated {len(updated_orders)} existing order(s).")

    if new_count == 0:
        log("  No new orders since last check.")
    return new_count


def main():
    if not HAS_OPENPYXL:
        print("[ERROR] openpyxl required. Install: pip install openpyxl")
        sys.exit(1)

    import argparse
    parser = argparse.ArgumentParser(description="3-Month Plan Order Monitor v1.1")
    parser.add_argument("--once", action="store_true", help="Single check and exit")
    parser.add_argument("--interval", type=int, default=POLL_INTERVAL, help=f"Poll interval in seconds (default: {POLL_INTERVAL})")
    parser.add_argument("--output", type=str, default="three_month_orders.xlsx", help="Output Excel file path")
    parser.add_argument("--no-slack", action="store_true", help="Disable Slack notifications")
    args = parser.parse_args()

    output_file = Path(args.output)
    if not output_file.exists():
        create_workbook(output_file)

    print("=" * 60)
    print("  📦 3-MONTH PLAN ORDER MONITOR v1.1")
    print("=" * 60)
    print(f"  Watching:    {', '.join(WATCH_PRODUCTS)}")
    print(f"  Output:      {output_file.resolve()}")
    print(f"  Interval:    {args.interval}s ({args.interval // 60} min)")
    print(f"  Slack:       {'Disabled' if args.no_slack else 'Enabled'}")
    print(f"  Databricks:  {DATABRICKS['host']}")
    print(f"  Catalog:     {DATABRICKS['catalog']}.{DATABRICKS['schema']}")
    print("=" * 60)

    if args.once:
        new = poll(output_file, send_alerts=not args.no_slack)
        print(f"\n{'✅' if new else 'ℹ️'}  Done. {new} new order(s) recorded to {output_file}")
        sys.exit(0)

    log("Starting continuous monitoring... (Ctrl+C to stop)")
    total_new = 0
    try:
        while True:
            new = poll(output_file, send_alerts=not args.no_slack)
            total_new += new
            log(f"Next check in {args.interval}s... (Total new: {total_new})")
            print("-" * 60)
            time.sleep(args.interval)
    except KeyboardInterrupt:
        print(f"\n{'=' * 60}")
        log(f"Monitor stopped. Total new orders: {total_new}")
        log(f"Output: {output_file.resolve()}")
        print(f"{'=' * 60}")


if __name__ == "__main__":
    main()

